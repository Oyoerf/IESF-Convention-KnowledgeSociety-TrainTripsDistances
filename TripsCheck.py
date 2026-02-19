import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def ajouter_verifications_simple(fichier_entree, fichier_sortie=None):
    """
    Ajoute deux colonnes de vérification calculées en Python :
    1. Nombre de trajets pair/impair par personne
    2. Premier départ = dernière arrivée (trajet circulaire)
    """
    if fichier_sortie is None:
        fichier_sortie = fichier_entree.replace('.xlsx', '_vérifié.xlsx')
    
    # Lire avec pandas
    df = pd.read_excel(fichier_entree)
    
    # Calculer les vérifications
    verifications = []
    
    for idx, row in df.iterrows():
        nom = row.iloc[0]  # Première colonne = Nom
        
        # Filtrer toutes les lignes pour cette personne
        lignes_personne = df[df.iloc[:, 0] == nom]
        
        # Vérification 1: Nombre de trajets pair?
        nb_trajets = len(lignes_personne)
        est_pair = "PAIR" if nb_trajets % 2 == 0 else "IMPAIR"
        
        # Vérification 2: Premier départ = dernière arrivée?
        if len(lignes_personne) > 0:
            premier_depart = lignes_personne.iloc[0, 3]  # Colonne D (index 3)
            derniere_arrivee = lignes_personne.iloc[-1, 4]  # Colonne E (index 4)
            est_circulaire = "OUI" if premier_depart == derniere_arrivee else "NON"
        else:
            est_circulaire = "N/A"
        
        verifications.append({
            'Nb trajets pair?': est_pair,
            'Trajet circulaire?': est_circulaire
        })
    
    # Ajouter les colonnes au DataFrame
    df_verif = pd.DataFrame(verifications)
    df_final = pd.concat([df, df_verif], axis=1)
    
    # Sauvegarder
    df_final.to_excel(fichier_sortie, index=False)
    
    # Formater avec openpyxl
    wb = load_workbook(fichier_sortie)
    sheet = wb.active
    
    # Trouver colonnes de vérification (dernières colonnes)
    derniere_col = sheet.max_column
    
    # Formater headers des colonnes de vérification
    for col in [derniere_col - 1, derniere_col]:
        cell = sheet.cell(1, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Colorer les cellules selon résultat
    for row in range(2, sheet.max_row + 1):
        # Colonne "Nb trajets pair?"
        cell_pair = sheet.cell(row, derniere_col - 1)
        if cell_pair.value == "IMPAIR":
            cell_pair.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Colonne "Trajet circulaire?"
        cell_circ = sheet.cell(row, derniere_col)
        if cell_circ.value == "NON":
            cell_circ.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif cell_circ.value == "OUI":
            cell_circ.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    wb.save(fichier_sortie)
    
    # Statistiques
    nb_impair = sum(1 for v in verifications if v['Nb trajets pair?'] == 'IMPAIR')
    nb_non_circulaire = sum(1 for v in verifications if v['Trajet circulaire?'] == 'NON')
    
    print(f"\nVérifications ajoutées : {fichier_sortie}")
    print(f"  - Personnes avec nombre impair de trajets: {nb_impair}")
    print(f"  - Trajets non circulaires: {nb_non_circulaire}")

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        ajouter_verifications_simple(sys.argv[1])
    else:
        ajouter_verifications_simple('trajets.xlsx')